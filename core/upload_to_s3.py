import os
import boto3
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from datetime import datetime
import re
import sys
from aws_config import AWS_ACCESS_KEY, AWS_SECRET_KEY, AWS_REGION

# Constants
ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png'}
MAX_FILE_SIZE_MB = 5
EXCEL_FILE = 'students.xlsx'
BUCKET_NAME = 'ict-attendance'

# Initialize S3 client
s3 = boto3.client(
    's3',
    region_name=AWS_REGION,
    aws_access_key_id=AWS_ACCESS_KEY,
    aws_secret_access_key=AWS_SECRET_KEY
)


def allowed_file(filename):
    """Check allowed file extension."""
    _, ext = os.path.splitext(filename)
    return ext.lower() in ALLOWED_EXTENSIONS


def file_size_okay(file_obj):
    """Check if file is within max size."""
    file_obj.seek(0, os.SEEK_END)
    size_mb = file_obj.tell() / (1024 * 1024)
    file_obj.seek(0)
    return size_mb <= MAX_FILE_SIZE_MB


def upload_file_to_s3(bucket_name, file_path, s3_key):
    """Upload a local file to S3 with the object key."""
    try:
        print(f"Uploading to S3 → Bucket: {bucket_name}, Key: {s3_key}")
        s3.upload_file(file_path, bucket_name, s3_key)
    except Exception as e:
        raise Exception(f"Upload failed: {e}")


def sanitize_for_s3_key(text: str) -> str:
    """Remove unsafe characters and replace spaces with underscores."""
    text = text.strip().replace(" ", "_")
    text = re.sub(r'[^a-zA-Z0-9_\-]', '', text)
    return text


def update_student_excel_and_upload(batch_name, er_number, name):
    """Update Excel, one sheet per batch, and upload to S3 root."""
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    if batch_name in wb.sheetnames:
        sheet = wb[batch_name]
    else:
        sheet = wb.create_sheet(batch_name)
        sheet.append(["ER Number", "Student Name", "Upload Date & Time"])

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet.append([er_number, name, now])
    wb.save(EXCEL_FILE)

    try:
        excel_key = f"students/{batch_name}/{subject_code}.xlsx"
        upload_file_to_s3(BUCKET_NAME, EXCEL_FILE, excel_key)
    except Exception as e:
        raise Exception(f"Failed to upload Excel to S3: {e}")


def upload_multiple_images(batch_name, er_number, name, image_files):
    """Upload multiple images under batch folder prefix in S3 bucket."""
    er_number = er_number.strip()
    sanitized_batch_name = sanitize_for_s3_key(batch_name)
    sanitized_name = sanitize_for_s3_key(name)

    os.makedirs("uploads", exist_ok=True)
    upload_results = []

    for i, image_file in enumerate(image_files):
        filename = secure_filename(image_file.filename)
        extension = os.path.splitext(filename)[1].lower()

        if not allowed_file(filename):
            upload_results.append(f"Rejected {filename}: Invalid file type")
            continue

        if not file_size_okay(image_file):
            upload_results.append(f"Rejected {filename}: File too large (> {MAX_FILE_SIZE_MB} MB)")
            continue

        # Compose S3 key: <batch>/<er_number>_<name>_<index>.<ext>
        new_filename = f"{er_number}_{sanitized_name}_{i + 1}{extension}"
        s3_key = f"{sanitized_batch_name}/{new_filename}"
        local_path = os.path.join("uploads", new_filename)

        try:
            image_file.save(local_path)
            s3.upload_file(Filename=local_path, Bucket=BUCKET_NAME, Key=s3_key)
            upload_results.append(f"✅ Uploaded: {s3_key}")
        except Exception as e:
            upload_results.append(f"❌ Failed: {s3_key} -> {str(e)}")
        finally:
            if os.path.exists(local_path):
                os.remove(local_path)

    sys.dont_write_bytecode = True

    try:
        update_student_excel_and_upload(batch_name, er_number, name)
        upload_results.append("✅ Excel updated and uploaded to S3 root.")
    except Exception as e:
        upload_results.append(f"❌ Excel update failed: {e}")

    return upload_results


def mark_attendance_s3():
    """
    Dummy placeholder for attendance logic using images stored in S3.
    Replace with your real face-recognition logic as needed.
    """
    results = [
        "✅ S3: Student1.jpg matched and marked present",
        "❌ S3: Student2.jpg face not found"
    ]
    return results


if __name__ == '__main__':
    # CLI test stub (note: no file uploads possible here)
    batch_name = input("Enter Batch Name: ").strip()
    er_number = input("Enter Student ER Number: ").strip()
    student_name = input("Enter Student Name: ").strip()
    image_files = []  # For CLI testing, must be FileStorage objects for uploads

    results = upload_multiple_images(batch_name, er_number, student_name, image_files)
    for res in results:
        print(res)